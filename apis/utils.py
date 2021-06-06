from openpyxl import load_workbook
from .models import Book
from slugify import slugify

def validate_data(data):
    book = data
    error = []

    def convert_field(label, type):
        try:
            return type(book[label])
        except:
            error.append(f'{label} should be an integer')

    if book['price'] == 'None':
        book['price'] = 0
    else:
        book['price'] = convert_field('price', float)

    if book['num_reviews'] == 'None':
        book['num_reviews'] = 0
    else:
        book['num_reviews'] = convert_field('num_reviews', int)

    if book['rating'] == 'None':
        book['rating'] = 0
    else:
        book['rating'] = convert_field('rating', int)

    if book['thumbnail'] == 'None':
        book['thumbnail'] = ""

    if book['availability'] != 'True' or book['availability'] != 'False':
        if book['availability'] == 'None':
            book['availability'] = True

    if len(error) > 0:
        raise Exception(error)
    return book
        

def handle_file(file):
    wb = load_workbook(file)
    sheet = wb[wb.sheetnames[0]]
    raw_data = list()
    for row in sheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        raw_data.append(row_data)
    column = raw_data[0]
    data = raw_data[1:]

    new = 0
    update = 0
    error = 0
    error_msg = []
    for row in data:
        book = dict(zip(column, row))
        try:
            validated_book = validate_data(book)

            ## Check if current book exist
            has_current_book = Book.objects.filter(title = validated_book['title'])
            slug = slugify(validated_book['title'])
            print(slug)

            if has_current_book.count() > 0:
                has_current_book.update(
                    slug=slug,
                    title=validated_book['title'],
                    price=validated_book['price'],
                    description = validated_book['description'],
                    category = validated_book['category'],
                    thumbnail = validated_book['thumbnail'],
                    rating = validated_book['rating'],
                    num_reviews = validated_book['num_reviews'],
                    availablility = validated_book['availability'])
                update = update+1
            else:
                new_book = Book(
                    slug=slug,
                    title=validated_book['title'],
                    price=validated_book['price'],
                    description = validated_book['description'],
                    category = validated_book['category'],
                    thumbnail = validated_book['thumbnail'],
                    rating = validated_book['rating'],
                    num_reviews = validated_book['num_reviews'],
                    availablility = validated_book['availability'])
                new_book.save()
                new = new + 1
        except Exception as e: 
            error_msg.append(f"In book id: {book['id']}, {e}")
            error = error+1
    return { 'new': new, 'update': update, 'error': error, 'msg': error_msg }


def check_form(form):
    try:
        return save_record(form)
    except Exception as e: 
        return e
        

def save_record(record):
    has_current_book = Book.objects.filter(title = record['title'])
    slug = slugify(record['title'])
    if has_current_book.count() > 0:
        has_current_book.update(
                    slug=slug,
                    title=record['title'],
                    price=record['price'],
                    description = record['description'],
                    category = record['category'],
                    thumbnail = record['thumbnail'],
                    rating = record['rating'],
                    num_reviews = record['num_reviews'],
                    availablility = record['availability'])
        return f"{record['title']} is updated in the database"
    else:
        new_book = Book(
                    slug=slug,
                    title=record['title'],
                    price=record['price'],
                    description = record['description'],
                    category = record['category'],
                    thumbnail = record['thumbnail'],
                    rating = record['rating'],
                    num_reviews = record['num_reviews'],
                    availablility = record['availability'])
        new_book.save()
        return f"{record['title']} is added to the database"